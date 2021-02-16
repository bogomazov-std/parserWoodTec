import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# div catalog-section
# div catalog item-views table
# div col-md-4 col-sm-6 col-xs-12
# div item
# span label-right

URL_MAIN = "https://woodtec.com.ru"
URL = "https://woodtec.com.ru/catalog/"
HEADERS = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0Win64x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.150 Safari/537.36',
    'accept': '*/*'
}


def getHtml(url, params=None):
    r = requests.get(url, params)
    return r


def getInfoCatalog(item):
    soup = BeautifulSoup(str(item), 'html.parser')
    link = soup.find('a', class_='catalog-section__name')
    title = link.get_text()
    urlCatalog = link.get('href')
    return {
        'title': title,
        'url': urlCatalog
    }


def getCatalog(htmlText):
    allCatalog = []
    soup = BeautifulSoup(htmlText, 'html.parser')
    items = soup.find_all('div', class_='catalog-section')
    if items:
        for item in items:
            objCatalog = getInfoCatalog(item)
            allCatalog.append(objCatalog)
        return allCatalog
    else:
        return None


def getAllCatalog(URL, itemUrl=''):
    url = URL + itemUrl
    html = getHtml(url, HEADERS)
    if html.status_code == 200:
        return getCatalog(html.text)



def getProduct(url, percent):
    html = getHtml(URL_MAIN+url)
    soup = BeautifulSoup(html.text, 'html.parser')
    title = soup.find('h1').get_text()
    priceOld = soup.find('span', class_='price__rub_old')
    if(priceOld):
        priceRUB = soup.find('span', class_='price__rub_old').get_text()
        priceUSD = soup.find('span', class_='price__usd_old').get_text()
    else:
        priceRUB = soup.find('span', class_='price__rub').get_text()
        priceUSD = soup.find('span', class_='price__usd').get_text()
    purchasingPrice = priceRUB.replace('₽', '')
    purchasingPrice = purchasingPrice.replace(' ', '')
    purchasingPrice = float(purchasingPrice) * percent
    return {
        'title': title,
        'priceUSD': priceUSD,
        'priceRUB': priceRUB,
        'purchasingPrice': purchasingPrice
    }



def getUrlProducts(percent, url, headers=HEADERS):
    allProducts = []
    html = getHtml(url, headers)
    soup = BeautifulSoup(html.text, 'html.parser')
    productsCard = soup.find('div', class_='catalog item-views table')
    products = productsCard.find_all(
        'div', class_='col-md-4 col-sm-6 col-xs-12')
    for product in products:
        title = product.find('div', class_='title')
        linkProduct = title.find('a')
        urlProduct = linkProduct.get('href')
        allProducts.append(
            getProduct(urlProduct, percent)
        )
    return allProducts



def getAllProducts(allCatalog):
    for item in allCatalog:
        percent = 0.75
        if item['title'] == 'Инструмент':
            percent = 0.7
        if item['addition']:
            for itemAddition in item['addition']:
                url = URL_MAIN + itemAddition['url']
                html = getHtml(url, HEADERS)
                soup = BeautifulSoup(html.text, 'html.parser')
                page = soup.find('ul', class_='pagination')
                if not page:
                    itemAddition['product'] = getUrlProducts(percent, url)
                else:
                    pages = len(page.find_all('li'))
                    allPage = []
                    for number in range(1, pages):
                        headers = HEADERS
                        headers['PPAGEN_2'] = '1'
                        headers['PAGEN_1'] = str(number)
                        for product in getUrlProducts(percent, url, headers):
                            allPage.append(product)
                    itemAddition['product'] = allPage
        else:
            url = URL_MAIN + item['url']
            item['product'] = getUrlProducts(percent, url)
    return allCatalog



def pullColumnXl(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value



def writeXl(allProduct):
    wb = Workbook()
    wb.active.title = 'info'
    for i in allProduct:
        ws = wb.create_sheet(i['title'].replace('Фрезерно-гравировальные', 'ФГ'))
        if i['addition']:
            for product in i['addition']:
                ws.append([product['title']])
                redFill = PatternFill(start_color='FFB1FF',
                                      end_color='FFB1FF',
                                      fill_type='solid')
                blueFill = PatternFill(start_color='9FC5E8',
                                      end_color='9FC5E8',
                                      fill_type='solid')
                ws['A{}'.format(ws.max_row)].fill = redFill
                ws.append(['Наименование', 'Цена в долларах', 'Цена в рублях', 'Закупочная'])
                for col in 'ABCD':
                    ws[col+str(ws.max_row)].fill = blueFill
                #color.font = PatternFill(bgColor="ffa6ed", fill_type="solid")
                for gg in product['product']:
                    ws.append([gg['title'], gg['priceUSD'], gg['priceRUB'], gg['purchasingPrice']])
        else:
            for k in i['product']:
                ws.append([k['title'], k['priceUSD'], k['priceRUB'], k['purchasingPrice']])
        pullColumnXl(ws)
    wb.save("sample.xlsx")


def parse():
    allCatalog = getAllCatalog(URL)
    for item in allCatalog:
        item['addition'] = getAllCatalog(URL_MAIN, item['url'])
    allProduct = getAllProducts(allCatalog)
    writeXl(allProduct)

parse()

